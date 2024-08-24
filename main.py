from openpyxl import Workbook
wb = Workbook()

ws = wb.active
ws["A1"] = "Dulce1"

ws.append([1,2,3])

import datetime
ws["A2"] = datetime.datetime.now()
wb.save("Sample.xlsx")
